from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse
from fastapi.exceptions import RequestValidationError
from sqlalchemy.orm import Session
from typing import List, Optional
import shutil
import os
from datetime import datetime
import pandas as pd
from io import BytesIO
import sqlite3

from . import models, schemas, crud, migration
from .database import engine, get_db

# Create database tables
models.Base.metadata.create_all(bind=engine)

# Run migrations to add any new columns to existing databases
try:
    migration.migrate_database()
except Exception as e:
    print(f"Migration warning: {e}")

app = FastAPI(title="Job Application Manager")

# Add validation exception handler to log errors
@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    print(f"Validation Error: {exc.errors()}")
    print(f"Request body: {await request.body()}")
    return JSONResponse(
        status_code=422,
        content={"detail": exc.errors()},
    )

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Create upload directories
os.makedirs("uploads/cvs", exist_ok=True)
os.makedirs("uploads/coverletters", exist_ok=True)

@app.get("/")
def read_root():
    return {"message": "Job Application Manager API", "status": "running"}

@app.post("/applications/", response_model=schemas.JobApplication)
def create_application(
    application: schemas.JobApplicationCreate,
    db: Session = Depends(get_db)
):
    try:
        print(f"Received application data: {application.model_dump()}")
        return crud.create_application(db=db, application=application)
    except Exception as e:
        print(f"Error creating application: {str(e)}")
        raise

@app.get("/applications/", response_model=List[schemas.JobApplication])
def read_applications(
    skip: int = 0,
    limit: int = 1000,
    status: Optional[str] = None,
    domain: Optional[str] = None,
    search: Optional[str] = None,
    work_type: Optional[str] = None,
    tags: Optional[str] = None,
    include_archived: bool = False,
    sort_by: Optional[str] = "created_at",
    sort_order: Optional[str] = "desc",
    status_stage: Optional[str] = None,
    db: Session = Depends(get_db)
):
    applications = crud.get_applications(
        db, skip=skip, limit=limit, status=status, domain=domain,
        search=search, work_type=work_type, tags=tags,
        include_archived=include_archived, sort_by=sort_by, sort_order=sort_order,
        status_stage=status_stage
    )
    return applications

@app.get("/applications/{application_id}", response_model=schemas.JobApplication)
def read_application(application_id: int, db: Session = Depends(get_db)):
    db_application = crud.get_application(db, application_id=application_id)
    if db_application is None:
        raise HTTPException(status_code=404, detail="Application not found")
    return db_application

@app.put("/applications/{application_id}", response_model=schemas.JobApplication)
def update_application(
    application_id: int,
    application: schemas.JobApplicationUpdate,
    db: Session = Depends(get_db)
):
    db_application = crud.update_application(db, application_id=application_id, application=application)
    if db_application is None:
        raise HTTPException(status_code=404, detail="Application not found")
    return db_application

@app.delete("/applications/{application_id}")
def delete_application(application_id: int, db: Session = Depends(get_db)):
    success = crud.delete_application(db, application_id=application_id)
    if not success:
        raise HTTPException(status_code=404, detail="Application not found")
    return {"message": "Application deleted successfully"}

# Archive/Unarchive endpoints
@app.put("/applications/{application_id}/archive")
def archive_application(application_id: int, db: Session = Depends(get_db)):
    db_application = crud.archive_application(db, application_id=application_id, archive=True)
    if db_application is None:
        raise HTTPException(status_code=404, detail="Application not found")
    return {"message": "Application archived successfully"}

@app.put("/applications/{application_id}/unarchive")
def unarchive_application(application_id: int, db: Session = Depends(get_db)):
    db_application = crud.archive_application(db, application_id=application_id, archive=False)
    if db_application is None:
        raise HTTPException(status_code=404, detail="Application not found")
    return {"message": "Application unarchived successfully"}

# Bulk operations
@app.post("/applications/bulk/delete")
def bulk_delete_applications(application_ids: List[int], db: Session = Depends(get_db)):
    count = crud.bulk_delete_applications(db, application_ids=application_ids)
    return {"message": f"Deleted {count} applications successfully"}

@app.post("/applications/bulk/archive")
def bulk_archive_applications(application_ids: List[int], db: Session = Depends(get_db)):
    count = crud.bulk_archive_applications(db, application_ids=application_ids, archive=True)
    return {"message": f"Archived {count} applications successfully"}

@app.post("/applications/bulk/unarchive")
def bulk_unarchive_applications(application_ids: List[int], db: Session = Depends(get_db)):
    count = crud.bulk_archive_applications(db, application_ids=application_ids, archive=False)
    return {"message": f"Unarchived {count} applications successfully"}

@app.post("/applications/bulk/update-status")
def bulk_update_status(application_ids: List[int], status: str, stage: Optional[str] = None, db: Session = Depends(get_db)):
    count = crud.bulk_update_status(db, application_ids=application_ids, status=status, stage=stage)
    return {"message": f"Updated status for {count} applications successfully"}

# Get all unique tags
@app.get("/tags/")
def get_all_tags(db: Session = Depends(get_db)):
    tags = crud.get_all_tags(db)
    return tags

@app.post("/applications/{application_id}/upload/{doc_type}")
async def upload_document(
    application_id: int,
    doc_type: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    if doc_type not in ["cv", "coverletter"]:
        raise HTTPException(status_code=400, detail="Invalid document type")
    
    # Check if application exists
    db_application = crud.get_application(db, application_id=application_id)
    if db_application is None:
        raise HTTPException(status_code=404, detail="Application not found")
    
    # Save file
    file_extension = os.path.splitext(file.filename)[1]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{application_id}_{timestamp}{file_extension}"
    
    if doc_type == "cv":
        filepath = f"uploads/cvs/{filename}"
    else:
        filepath = f"uploads/coverletters/{filename}"
    
    with open(filepath, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Update database
    updated_app = crud.update_document(db, application_id, doc_type, file.filename, filepath)
    
    return {"message": f"{doc_type} uploaded successfully", "filename": file.filename}

@app.get("/applications/{application_id}/download/{doc_type}")
def download_document(
    application_id: int,
    doc_type: str,
    db: Session = Depends(get_db)
):
    db_application = crud.get_application(db, application_id=application_id)
    if db_application is None:
        raise HTTPException(status_code=404, detail="Application not found")
    
    if doc_type == "cv":
        filepath = db_application.cv_filepath
        filename = db_application.cv_filename
    elif doc_type == "coverletter":
        filepath = db_application.coverletter_filepath
        filename = db_application.coverletter_filename
    else:
        raise HTTPException(status_code=400, detail="Invalid document type")
    
    if not filepath or not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(filepath, filename=filename)

@app.get("/statistics/", response_model=schemas.ApplicationStats)
def get_statistics(db: Session = Depends(get_db)):
    return crud.get_statistics(db)

@app.get("/export/excel")
def export_to_excel(db: Session = Depends(get_db)):
    """Export all job applications and cold messages to a multi-sheet Excel file"""
    import json as _json
    from openpyxl.styles import PatternFill, Font, Alignment

    applications = crud.get_all_applications_for_export(db)
    cold_messages = crud.get_cold_messages(db)

    # ── Sheet 1: Job Applications ────────────────────────────────────────
    apps_data = []
    for app in applications:
        # Derive latest rejection stage from status_history
        rejection_stage = ""
        try:
            history = app.status_history if isinstance(app.status_history, list) else _json.loads(app.status_history or "[]")
            for entry in reversed(history):
                if entry.get("stage"):
                    rejection_stage = entry["stage"]
                    break
        except Exception:
            pass

        tags_raw = app.tags
        if isinstance(tags_raw, str):
            try:
                tags_raw = _json.loads(tags_raw)
            except Exception:
                tags_raw = []
        tags_str = ", ".join(tags_raw) if tags_raw else ""

        # Networking contacts (JSON list of {name, linkedin_url, ...})
        net_contacts = app.networking_contacts or []
        if isinstance(net_contacts, str):
            try:
                net_contacts = _json.loads(net_contacts)
            except Exception:
                net_contacts = []
        net_contacts_str = "; ".join(
            f"{c.get('name', '')} ({c.get('linkedin_url', c.get('email', ''))})".strip("()")
            for c in net_contacts if isinstance(c, dict)
        ) if net_contacts else ""

        apps_data.append({
            "ID": app.id,
            "Company Name": app.company_name,
            "Job Title": app.job_title,
            "Domain": app.domain or "",
            "Location": app.location or "",
            "Work Type": app.work_type or "",
            "Applied On (Portal)": app.applied_on or "",
            "Status": app.status,
            "Rejection Stage": rejection_stage,
            "Application Date": app.application_date.strftime("%Y-%m-%d") if app.application_date else "",
            "Application Deadline": app.application_deadline.strftime("%Y-%m-%d") if app.application_deadline else "",
            "Interview Date": app.interview_date.strftime("%Y-%m-%d") if app.interview_date else "",
            "Salary Min": app.salary_min or "",
            "Salary Max": app.salary_max or "",
            "Tags": tags_str,
            "Job URL": app.job_url or "",
            "Contact Person": app.contact_person or "",
            "Contact Email": app.contact_email or "",
            "Contact LinkedIn": app.contact_linkedin or "",
            "Cold Outreach Sent": "Yes" if app.contact_cold_message_sent else "No",
            "Cold Outreach Via": app.contact_cold_message_via or "",
            "Cold Contact Category": app.contact_cold_contact_category or "",
            "Cold Contact Email": app.contact_cold_contact_email or "",
            "Cold Message Body": app.contact_cold_message_body or "",
            "Networking Contacts": net_contacts_str,
            "References": app.references or "",
            "CV Uploaded": "Yes" if app.cv_filename else "No",
            "Cover Letter Uploaded": "Yes" if app.coverletter_filename else "No",
            "Interview Notes": app.interview_notes or "",
            "Interview Questions": app.interview_questions or "",
            "Notes": app.notes or "",
            "Job Description": app.job_description or "",
            "Archived": "Yes" if app.is_archived else "No",
            "Created At": app.created_at.strftime("%Y-%m-%d") if app.created_at else "",
            "Updated At": app.updated_at.strftime("%Y-%m-%d") if app.updated_at else "",
        })

    # ── Sheet 2: Cold Messages ───────────────────────────────────────────
    cm_data = []
    for msg in cold_messages:
        cm_data.append({
            "ID": msg.id,
            "Contact Name": msg.contact_name,
            "Company": msg.company_name or "",
            "Contact Email": msg.contact_email or "",
            "Contact LinkedIn": msg.contact_linkedin or "",
            "Via": msg.via,
            "Category": msg.category or "",
            "Subject": msg.subject or "",
            "Message Body": msg.message_body or "",
            "Sent Date": msg.sent_date.strftime("%Y-%m-%d") if msg.sent_date else "",
            "Got Reply": "Yes" if msg.got_reply else "No",
            "Linked Connection ID": msg.connection_id or "",
            "Notes": msg.notes or "",
            "Created At": msg.created_at.strftime("%Y-%m-%d") if msg.created_at else "",
            "Updated At": msg.updated_at.strftime("%Y-%m-%d") if msg.updated_at else "",
        })

    # ── Build Excel ──────────────────────────────────────────────────────
    output = BytesIO()
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def style_sheet(ws, df):
        for idx, col in enumerate(df.columns):
            col_letter = ws.cell(row=1, column=idx + 1).column_letter
            cell = ws.cell(row=1, column=idx + 1)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            max_len = max(df[col].astype(str).apply(len).max() if not df.empty else 0, len(col))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 60)
        ws.row_dimensions[1].height = 20

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_apps = pd.DataFrame(apps_data) if apps_data else pd.DataFrame(columns=list(apps_data[0].keys()) if apps_data else [])
        df_apps.to_excel(writer, sheet_name='Job Applications', index=False)
        style_sheet(writer.sheets['Job Applications'], df_apps)

        df_cm = pd.DataFrame(cm_data) if cm_data else pd.DataFrame(columns=["ID","Contact Name","Company","Contact Email","Contact LinkedIn","Via","Category","Subject","Message Body","Sent Date","Got Reply","Linked Connection ID","Notes","Created At","Updated At"])
        df_cm.to_excel(writer, sheet_name='Cold Messages', index=False)
        style_sheet(writer.sheets['Cold Messages'], df_cm)

        # ── Sheet 3: LinkedIn Connections ────────────────────────────────────
        connections = crud.get_all_connections_for_export(db)
        conn_data = [
            {
                "ID": c.id,
                "Contact Name": c.contact_name,
                "Company": c.company_name or "",
                "LinkedIn Profile": c.linkedin_profile_id or "",
                "Category": c.category or "",
                "Stage": c.stage or "Requested",
                "Status": c.connection_status,
                "Cold Message Sent": "Yes" if c.cold_message_sent else "No",
                "Linked Cold Message ID": c.cold_message_id or "",
                "Requested On": c.requested_on.strftime("%Y-%m-%d") if c.requested_on else "",
                "Accepted On": c.accepted_on.strftime("%Y-%m-%d") if c.accepted_on else "",
                "Follow Up Date": c.follow_up_date.strftime("%Y-%m-%d") if c.follow_up_date else "",
                "Notes": c.notes or "",
                "Created At": c.created_at.strftime("%Y-%m-%d") if c.created_at else "",
                "Updated At": c.updated_at.strftime("%Y-%m-%d") if c.updated_at else "",
            }
            for c in connections
        ]
        conn_cols = ["ID","Contact Name","Company","LinkedIn Profile","Category","Stage","Status","Cold Message Sent","Linked Cold Message ID","Requested On","Accepted On","Follow Up Date","Notes","Created At","Updated At"]
        df_conn = pd.DataFrame(conn_data) if conn_data else pd.DataFrame(columns=conn_cols)
        df_conn.to_excel(writer, sheet_name='LinkedIn Connections', index=False)
        style_sheet(writer.sheets['LinkedIn Connections'], df_conn)

    output.seek(0)
    filename = f"job_tracker_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/domains/")
def get_unique_domains(db: Session = Depends(get_db)):
    """Get list of unique domains from applications"""
    domains = db.query(models.JobApplication.domain).filter(
        models.JobApplication.domain.isnot(None)
    ).distinct().all()
    return [domain[0] for domain in domains if domain[0]]


# ─── Cold Message Endpoints ──────────────────────────────────────────

@app.get("/cold-messages/statistics/", response_model=schemas.ColdMessageStats)
def get_cold_message_stats(db: Session = Depends(get_db)):
    return crud.get_cold_message_stats(db)

@app.post("/cold-messages/", response_model=schemas.ColdMessage)
def create_cold_message(msg: schemas.ColdMessageCreate, db: Session = Depends(get_db)):
    return crud.create_cold_message(db, msg)

@app.get("/cold-messages/", response_model=List[schemas.ColdMessage])
def list_cold_messages(
    search: Optional[str] = None,
    via: Optional[str] = None,
    category: Optional[str] = None,
    sort_by: Optional[str] = "created_at",
    sort_order: Optional[str] = "desc",
    db: Session = Depends(get_db),
):
    return crud.get_cold_messages(db, search=search, via=via, category=category,
                                  sort_by=sort_by, sort_order=sort_order)

@app.get("/cold-messages/{msg_id}", response_model=schemas.ColdMessage)
def get_cold_message(msg_id: int, db: Session = Depends(get_db)):
    msg = crud.get_cold_message(db, msg_id)
    if not msg:
        raise HTTPException(status_code=404, detail="Cold message not found")
    return msg

@app.put("/cold-messages/{msg_id}", response_model=schemas.ColdMessage)
def update_cold_message(msg_id: int, data: schemas.ColdMessageUpdate, db: Session = Depends(get_db)):
    msg = crud.update_cold_message(db, msg_id, data)
    if not msg:
        raise HTTPException(status_code=404, detail="Cold message not found")
    return msg

@app.delete("/cold-messages/{msg_id}")
def delete_cold_message(msg_id: int, db: Session = Depends(get_db)):
    if not crud.delete_cold_message(db, msg_id):
        raise HTTPException(status_code=404, detail="Cold message not found")
    return {"message": "Cold message deleted successfully"}


# ─── LinkedIn Connection Endpoints ───────────────────────────────────

@app.get("/connections/statistics/", response_model=schemas.LinkedInConnectionStats)
def get_connection_stats(db: Session = Depends(get_db)):
    return crud.get_connection_stats(db)

@app.post("/connections/", response_model=schemas.LinkedInConnection)
def create_connection(data: schemas.LinkedInConnectionCreate, db: Session = Depends(get_db)):
    return crud.create_connection(db, data)


@app.get("/connections/check-duplicate")
def check_connection_duplicate(
    name: str,
    company: Optional[str] = None,
    linkedin_url: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Check if a connection already exists by name + company + LinkedIn URL."""
    existing = crud.check_connection_duplicate(
        db,
        contact_name=name,
        company_name=company or None,
        linkedin_profile_id=linkedin_url or None,
    )
    if existing:
        return {"exists": True, "id": existing.id}
    return {"exists": False}

@app.get("/connections/", response_model=List[schemas.LinkedInConnection])
def list_connections(
    search: Optional[str] = None,
    status: Optional[str] = None,
    category: Optional[str] = None,
    cold_message_sent: Optional[bool] = None,
    stage: Optional[str] = None,
    sort_by: Optional[str] = "created_at",
    sort_order: Optional[str] = "desc",
    db: Session = Depends(get_db),
):
    return crud.get_connections(
        db, search=search, status=status, category=category,
        cold_message_sent=cold_message_sent, stage=stage, sort_by=sort_by, sort_order=sort_order
    )

@app.get("/connections/{conn_id}", response_model=schemas.LinkedInConnection)
def get_connection(conn_id: int, db: Session = Depends(get_db)):
    conn = crud.get_connection(db, conn_id)
    if not conn:
        raise HTTPException(status_code=404, detail="Connection not found")
    return conn

@app.put("/connections/{conn_id}", response_model=schemas.LinkedInConnection)
def update_connection(conn_id: int, data: schemas.LinkedInConnectionUpdate, db: Session = Depends(get_db)):
    conn = crud.update_connection(db, conn_id, data)
    if not conn:
        raise HTTPException(status_code=404, detail="Connection not found")
    return conn

@app.delete("/connections/{conn_id}")
def delete_connection(conn_id: int, db: Session = Depends(get_db)):
    if not crud.delete_connection(db, conn_id):
        raise HTTPException(status_code=404, detail="Connection not found")
    return {"message": "Connection deleted successfully"}


# Backup and Restore endpoints
@app.get("/api/backup")
def backup_database():
    """Download the current database file backup"""
    db_path = "job_tracker.db"
    
    if not os.path.exists(db_path):
        raise HTTPException(status_code=404, detail="Database file not found")
        
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"job_tracker_backup_{timestamp}.db"
    
    return FileResponse(
        path=db_path,
        filename=filename,
        media_type="application/octet-stream"
    )

@app.post("/api/restore/preview")
async def preview_restore(file: UploadFile = File(...)):
    """Compare a backup .db file against the current database and return a diff summary."""
    import tempfile, os as _os

    if not file.filename.endswith('.db'):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload a .db file")

    # Save uploaded file to a temp location
    with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp:
        tmp_path = tmp.name
        shutil.copyfileobj(file.file, tmp)

    try:
        # Validate the backup file
        try:
            backup_conn = sqlite3.connect(f"file:{tmp_path}?mode=ro", uri=True)
            cursor = backup_conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='job_applications'")
            if not cursor.fetchone():
                raise ValueError("Not a valid Job Application Manager database (missing job_applications table).")
        except sqlite3.DatabaseError:
            raise ValueError("The uploaded file is not a valid SQLite database.")

        # Read backup applications
        cursor.execute("""
            SELECT id, company_name, job_title, status, application_date
            FROM job_applications
        """)
        backup_rows = cursor.fetchall()
        backup_conn.close()

        # Read current applications
        current_db_path = "job_tracker.db"
        current_conn = sqlite3.connect(f"file:{current_db_path}?mode=ro", uri=True)
        cur2 = current_conn.cursor()
        cur2.execute("""
            SELECT id, company_name, job_title, status, application_date
            FROM job_applications
        """)
        current_rows = cur2.fetchall()
        current_conn.close()

        # Build lookup sets by id
        backup_by_id = {row[0]: row for row in backup_rows}
        current_by_id = {row[0]: row for row in current_rows}

        backup_ids = set(backup_by_id.keys())
        current_ids = set(current_by_id.keys())

        # Applications in backup but not in current = will be "added back"
        added_ids = backup_ids - current_ids
        # Applications in current but not in backup = will be removed
        removed_ids = current_ids - backup_ids
        # Applications in both
        unchanged_ids = backup_ids & current_ids

        def row_to_dict(row):
            return {
                "id": row[0],
                "company_name": row[1],
                "job_title": row[2],
                "status": row[3],
                "application_date": row[4],
            }

        return {
            "current_count": len(current_rows),
            "backup_count": len(backup_rows),
            "to_add": len(added_ids),
            "to_remove": len(removed_ids),
            "unchanged": len(unchanged_ids),
            "added_items": [row_to_dict(backup_by_id[i]) for i in sorted(added_ids)],
            "removed_items": [row_to_dict(current_by_id[i]) for i in sorted(removed_ids)],
        }

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Preview failed: {str(e)}")
    finally:
        _os.unlink(tmp_path)


@app.post("/api/restore")

async def restore_database(file: UploadFile = File(...)):
    """Restore database from uploaded file with validation"""
    if not file.filename.endswith('.db'):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload a .db file")
    
    db_path = "job_tracker.db"
    temp_path = f"temp_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    backup_path = f"job_tracker.db.bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    try:
        # 1. Save uploaded file to a temporary location
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        # 2. Validate the database file
        try:
            conn = sqlite3.connect(temp_path)
            cursor = conn.cursor()
            
            # Check if it's a valid SQLite DB and has the main table
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='job_applications'")
            if not cursor.fetchone():
                conn.close()
                raise ValueError("The uploaded file does not appear to be a valid Job Application Manager database (missing job_applications table).")
                
            conn.close()
            
        except sqlite3.DatabaseError:
            raise ValueError("The uploaded file is not a valid SQLite database.")
            
        # 3. Create a backup of the current DB just in case
        if os.path.exists(db_path):
            shutil.copy2(db_path, backup_path)
            
        # CRITICAL FIX: Close all database connections before swapping the file
        # On Linux: Prevents reading from the old 'ghost' file handle
        # On Windows: Prevents PermissionError because the file is locked
        engine.dispose()
            
        # 4. Apply the new database
        try:
            shutil.move(temp_path, db_path)
        except PermissionError:
            # Retry mechanism for Windows if file is still briefly locked
            import time
            time.sleep(0.5)
            try:
                 shutil.move(temp_path, db_path)
            except Exception as e:
                # If we can't move it, try copy and delete
                shutil.copy2(temp_path, db_path)
                os.remove(temp_path)

        # 5. Run migration to ensure schema compatibility
        try:
            print("Running post-restore migration...")
            migration.migrate_database()
        except Exception as e:
            print(f"Post-restore migration failed: {e}")
            
        return {"message": "Database restored successfully. Please refresh the page."}
        
    except ValueError as e:
        # Invalid file - clean up temp and return error
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise HTTPException(status_code=400, detail=str(e))
        
    except Exception as e:
        # Unexpected error - try to restore from backup
        if os.path.exists(temp_path):
            os.remove(temp_path)
        if os.path.exists(backup_path) and os.path.exists(db_path):
            # If we messed up the main DB, try to restore
             shutil.copy2(backup_path, db_path)
             
        raise HTTPException(status_code=500, detail=f"Error restoring database: {str(e)}")
